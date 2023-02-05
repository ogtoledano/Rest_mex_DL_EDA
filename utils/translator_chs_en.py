import pandas as pd
import os
import torch
import nltk

import pandas as pd
from nltk.corpus import stopwords
from nltk.corpus import wordnet as wn
from sklearn.model_selection import train_test_split

MAX_LEN_SENTENCE=0
from transformers import T5Tokenizer,AutoTokenizer

from torch.utils.data import Dataset
from torch.utils.data import DataLoader

from nltk.stem import WordNetLemmatizer
import spacy
from spacy.lang.es.examples import sentences
import es_core_news_md
from nltk import SnowballStemmer
from transformers import MarianMTModel, MarianTokenizer
import nltk
from nltk.corpus import wordnet as wn


class CustomDataset(Dataset):
    def __init__(self, data):
        self.X = data['input_ids']
        self.attention = data['attention_mask']

    def __getitem__(self, idx):
        return {'input_ids': torch.tensor(self.X[idx]), 'attention_mask': torch.tensor(self.attention[idx])}

    def __len__(self):
        return len(self.X)


def removing_stop_words(texts):
    # Removing stop words
    stop_words = set(stopwords.words('spanish'))
    for i, text in enumerate(texts):
        tokens = nltk.word_tokenize(text)
        sentence = [word for word in tokens if word not in stop_words]
        texts[i] = ' '.join(sentence)


def lemmatization(texts):
    nlp = es_core_news_md.load()
    for i, text in enumerate(texts):
        doc = nlp(text)
        sentence = [word.lemma_.lower() for word in doc]
        texts[i] = ' '.join(sentence)


def normalize(texts):
    nlp = es_core_news_md.load()
    for i, text in enumerate(texts):
        doc = nlp(text)
        words = [t.orth_ for t in doc if not t.is_punct | t.is_stop]
        lexical_tokens = [t.lower() for t in words if len(t) > 3 and t.isalpha()]
        texts[i] = ' '.join(lexical_tokens)


def get_mt_model(translation, device):
    model_name = 'Helsinki-NLP/opus-mt-' + translation
    tokenizer = MarianTokenizer.from_pretrained(model_name)
    model = MarianMTModel.from_pretrained(model_name)
    return tokenizer, model


def format_batch_texts(language_code, batch_texts):
    formated_bach = [">>{}<< {}".format(language_code, ' '.join(nltk.word_tokenize(text[:500]))) for text in
                     batch_texts]
    return formated_bach


def perform_translation(batch_texts, device, model, tokenizer, language="es"):
    # Prepare the text data into appropriate format for the model
    formated_batch_texts = format_batch_texts(language, batch_texts)
    model.to(device)
    source_encoding = tokenizer(formated_batch_texts, return_tensors="pt", padding=True)

    encodings = {
        'input_ids': source_encoding['input_ids'],
        'attention_mask': source_encoding['attention_mask']
    }

    dataset = None

    if language == 'en':
        torch.save(encodings, "chs_mt5_translate_es.pt")
        dataset = CustomDataset(torch.load("chs_mt5_translate_es.pt"))
    else:
        torch.save(encodings, "chs_mt5_translate_en.pt")
        dataset = CustomDataset(torch.load("chs_mt5_translate_en.pt"))

    dataloader = DataLoader(dataset, batch_size=8, shuffle=True)

    # Generate translation using model
    translated = []

    for batch in dataloader:
        input_ids = batch['input_ids'].to(device)
        attention_mask = batch['attention_mask'].to(device)
        out = model.generate(input_ids=input_ids, attention_mask=attention_mask)
        translated += out

    # Convert the generated tokens indices back into text
    translated_texts = [tokenizer.decode(t, skip_special_tokens=True) for t in translated]

    return translated_texts


def combine_texts(original_texts, back_translated_batch):
    return set(original_texts + back_translated_batch)


def perform_back_translation_with_augmentation(batch_texts, device,first_model_tkn, first_model, second_model_tkn, second_model, original_language="es", temporary_language="en"):

    # Translate from Original to Temporary Language
    tmp_translated_batch = perform_translation(batch_texts, device, first_model, first_model_tkn, "en")

    # Translate Back to English
    back_translated_batch = perform_translation(tmp_translated_batch, device, second_model, second_model_tkn,
                                                original_language)

    # Return The Final Result
    # return combine_texts(original_texts, back_translated_batch)
    return back_translated_batch


def prepare_input(texts):
    for i, text in enumerate(texts):
        texts[i] = "multilabel classification: {} </s>".format(text)


def tokenize(data,tokenizer):

    source_encoding = tokenizer(
        data['input_text'],
        max_length=250,
        padding='max_length',
        pad_to_max_length=True,
        truncation=True,
    )

    targets_polarity = [value[0] for value in data['target_text']]

    target_encoding_polarity = tokenizer(
        targets_polarity,
        max_length=2,
        pad_to_max_length=False,
        truncation=False,
    )

    targets_polarity = [int(value) for value in targets_polarity]

    encodings = {
        'source_ids': source_encoding['input_ids'],
        'target_ids': target_encoding_polarity['input_ids'],
        'attention_mask': source_encoding['attention_mask'],
        'labels': torch.tensor(targets_polarity, dtype=torch.long)
    }

    return encodings


from pathlib import Path
from copy import copy
from typing import Union, Optional
import numpy as np
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def copy_excel_cell_range(
        src_ws: openpyxl.worksheet.worksheet.Worksheet,
        min_row: int = None,
        max_row: int = None,
        min_col: int = None,
        max_col: int = None,
        tgt_ws: openpyxl.worksheet.worksheet.Worksheet = None,
        tgt_min_row: int = 1,
        tgt_min_col: int = 1,
        with_style: bool = True
) -> openpyxl.worksheet.worksheet.Worksheet:

    if tgt_ws is None:
        tgt_ws = src_ws

    # https://stackoverflow.com/a/34838233/5741205
    for row in src_ws.iter_rows(min_row=min_row, max_row=max_row,
                                min_col=min_col, max_col=max_col):
        for cell in row:
            tgt_cell = tgt_ws.cell(
                row=cell.row + tgt_min_row - 1,
                column=cell.col_idx + tgt_min_col - 1,
                value=cell.value
            )
            if with_style and cell.has_style:
                # tgt_cell._style = copy(cell._style)
                tgt_cell.font = copy(cell.font)
                tgt_cell.border = copy(cell.border)
                tgt_cell.fill = copy(cell.fill)
                tgt_cell.number_format = copy(cell.number_format)
                tgt_cell.protection = copy(cell.protection)
                tgt_cell.alignment = copy(cell.alignment)
    return tgt_ws


def append_df_to_excel(
        filename: Union[str, Path],
        df: pd.DataFrame,
        sheet_name: str = 'Sheet1',
        startrow: Optional[int] = None,
        max_col_width: int = 30,
        autofilter: bool = False,
        fmt_int: str = "#,##0",
        fmt_float: str = "#,##0.00",
        fmt_date: str = "yyyy-mm-dd",
        fmt_datetime: str = "yyyy-mm-dd hh:mm",
        truncate_sheet: bool = False,
        storage_options: Optional[dict] = None,
        **to_excel_kwargs
) -> None:

    def set_column_format(ws, column_letter, fmt):
        for cell in ws[column_letter]:
            cell.number_format = fmt
    filename = Path(filename)
    file_exists = filename.is_file()
    # process parameters
    # calculate first column number
    # if the DF will be written using `index=True`, then `first_col = 2`, else `first_col = 1`
    first_col = int(to_excel_kwargs.get("index", True)) + 1
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    # save content of existing sheets
    if file_exists:
        wb = load_workbook(filename)
        sheet_names = wb.sheetnames
        sheet_exists = sheet_name in sheet_names
        sheets = {ws.title: ws for ws in wb.worksheets}

    with pd.ExcelWriter(
        filename.with_suffix(".xlsx"),
        engine="openpyxl",
        mode="a" if file_exists else "w",
        if_sheet_exists="new" if file_exists else None,
        date_format=fmt_date,
        datetime_format=fmt_datetime,
        storage_options=storage_options
    ) as writer:
        if file_exists:
            # try to open an existing workbook
            writer.book = wb
            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row
            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)
            # copy existing sheets
            writer.sheets = sheets
        else:
            # file doesn't exist, we are creating a new one
            startrow = 0

        # write out the DataFrame to an ExcelWriter
        df.to_excel(writer, sheet_name=sheet_name, **to_excel_kwargs)
        worksheet = writer.sheets[sheet_name]

        if autofilter:
            worksheet.auto_filter.ref = worksheet.dimensions

        for xl_col_no, dtyp in enumerate(df.dtypes, first_col):
            col_no = xl_col_no - first_col
            width = max(df.iloc[:, col_no].astype(str).str.len().max(),
                        len(df.columns[col_no]) + 6)
            width = min(max_col_width, width)
            column_letter = get_column_letter(xl_col_no)
            worksheet.column_dimensions[column_letter].width = width
            if np.issubdtype(dtyp, np.integer):
                set_column_format(worksheet, column_letter, fmt_int)
            if np.issubdtype(dtyp, np.floating):
                set_column_format(worksheet, column_letter, fmt_float)

    if file_exists and sheet_exists:
        # move (append) rows from new worksheet to the `sheet_name` worksheet
        wb = load_workbook(filename)
        # retrieve generated worksheet name
        new_sheet_name = set(wb.sheetnames) - set(sheet_names)
        if new_sheet_name:
            new_sheet_name = list(new_sheet_name)[0]
        # copy rows written by `df.to_excel(...)` to
        copy_excel_cell_range(
            src_ws=wb[new_sheet_name],
            tgt_ws=wb[sheet_name],
            tgt_min_row=startrow + 1,
            with_style=True
        )
        # remove new (generated by Pandas) worksheet
        del wb[new_sheet_name]
        wb.save(filename)
        wb.close()


def write_xlsx_file(reviews):
    df = pd.DataFrame(reviews, columns =['Title', 'Opinion', 'Class','Date'])

    if not os.path.exists('cuba_hotels_sentiment_en.xlsx'):
        append_df_to_excel('cuba_hotels_sentiment_en.xlsx', df, header=True, index=False,fmt_int='#')
    else:
        append_df_to_excel('cuba_hotels_sentiment_en.xlsx', df, header=None, index=False, fmt_int='#')


def build_dataset_and_dict():
    os.chdir('../')
    wdir = os.getcwd()
    device = "cuda:0" if torch.cuda.is_available() else "cpu"
    nltk.download('omw-1.4')
    nltk.download('wordnet')
    nltk.download('stopwords')
    nltk.download('punkt')
    file_train = wdir + "/datasets/cuba_hotels_sentiment.xlsx"
    df = pd.read_excel(r'' + file_train, engine='openpyxl')
    print(df.iloc[0, 1])
    print(df.iloc[350, 2])
    print(df.shape)

    reviews=[]

    tokenizer = T5Tokenizer.from_pretrained("google/mt5-small")
    dictionary_data = {'X': [], 'y': []}

    first_model_tkn, first_model = get_mt_model('es-en', device)
    second_model_tkn, second_model = get_mt_model('en-es', device)

    for i in range(df.shape[0]):
        title=perform_translation([str(df.iloc[i, 0])], device, first_model, first_model_tkn, "en")
        opinion=perform_translation([str(df.iloc[i, 1])], device, first_model, first_model_tkn, "en")
        reviews.append((title,opinion,str(df.iloc[i, 2]),str(df.iloc[i, 3])))

    write_xlsx_file(reviews)

if __name__ == "__main__":
    build_dataset_and_dict()

